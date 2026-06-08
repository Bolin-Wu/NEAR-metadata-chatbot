import os
import logging
from pathlib import Path
import tempfile
import re
from io import BytesIO
import time

import streamlit as st
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from langchain_chroma import Chroma
from langchain_openai import AzureOpenAIEmbeddings
from langchain_core.output_parsers import StrOutputParser
from langchain_core.documents import Document
from langchain_core.runnables import RunnablePassthrough
from langchain_groq import ChatGroq
from langchain_openai import ChatOpenAI  # For OpenAI-compatible endpoints
from langchain_core.prompts import PromptTemplate
from dotenv import load_dotenv
import shutil
import zipfile

load_dotenv()

# Configure logging (set to WARNING for production, INFO/DEBUG for development)
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

# ── Configuration ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NEAR Metadata Chatbot",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items=None
)

# ── Custom Styling (NEAR Brand) ───────────────────────────────────────────────
# Load external CSS file for clean separation of concerns
css_file = Path("styles/near_brand.css")
if css_file.exists():
    with open(css_file, "r") as f:
        css_content = f.read()
    st.markdown(f"<style>{css_content}</style>", unsafe_allow_html=True)
else:
    logger.warning(f"CSS file not found at {css_file}. Styling may not be applied.")

AZURE_EMBEDDING_API_VERSION = "2024-02-01"
EMBEDDING_MODEL_NAME = "text-embedding-3-small"
CHROMA_DB = "./chroma_azure_db"
CHROMA_ZIP_FILENAME = "chroma_azure_db.zip"
HUGGINGFACE_REPO_SECRET_NAME = "HUGGINGFACE_AZURE_REPO"
DATA_ROOT = "./data"

# Known databases (hardcoded as fallback when data/ folder not available)
KNOWN_DATABASES = [
    "Betula", "GAS_SNAC_S", "GENDER", "H70", "KP", 
    "OCTO-Twin", "SATSA", "SNAC-B", "SNAC-K", "SNAC-N", "SWEOLD", "TryBo", "ULSAM"
]

# LLM Model Names (for consistency across the app)
LLM_MODEL_GROQ = "Llama 3.1 8B (Groq free tier)"
LLM_MODEL_GPT = "GPT-5.4 Mini"
LLM_MODEL_XAI_GROK = "Grok 4.1 Fast Reasoning"

# LLM Model IDs (technical identifiers for API calls)
GROQ_MODEL_ID = "llama-3.1-8b-instant"
GPT_MODEL_ID = "gpt-5.4-mini"
XAI_GROK_MODEL_ID = "grok-4-1-fast-reasoning"
XAI_BASE_URL = "https://api.x.ai/v1/"

# LLM Hyperparameters
LLM_TEMPERATURE = 0.3           # Balanced: accurate answers with flexibility for general knowledge (0.0=deterministic, 1.0=creative)

# Rate Limiting (per browser session — protects Azure API credit from abuse)
RATE_LIMIT_MAX_GPT_QUERIES = 10      # Azure GPT calls allowed per session

# Retrieval Parameters
RETRIEVAL_K_BACKGROUND = 5      # Top-5 docs for cohort background context
RETRIEVAL_K_CONTEXT = 60        # Top-60 docs for variable definitions (increased for better category capture)
RETRIEVAL_K_CONTEXT_GROQ = 30   # Reduced for Groq free tier to stay within 6000 TPM token limit
EXACT_MATCH_LIMIT = 3          # Exact variable-name hits to merge ahead of semantic results
BACKGROUND_ENRICH_FALLBACK_DOCS = 3  # Fallback docs used when no exact variable match is found
BACKGROUND_ENRICH_MAX_DOCS = 3       # Max docs used to extract enrichment terms

# Safe way to get API keys
try:
    GROQ_API_KEY = st.secrets["GROQ_api_key"]
except (FileNotFoundError, KeyError, AttributeError):
    GROQ_API_KEY = os.getenv("GROQ_api_key")

try:
    AZURE_api_key = st.secrets["AZURE_api_key"]
except (FileNotFoundError, KeyError, AttributeError):
    AZURE_api_key = os.getenv("AZURE_api_key")

try:
    AZURE_openai_endpoint = st.secrets["AZURE_openai_endpoint"]
except (FileNotFoundError, KeyError, AttributeError):
    AZURE_openai_endpoint = os.getenv("AZURE_openai_endpoint")

try:
    AZURE_FOUNDRY_BASE_URL = st.secrets["AZURE_FOUNDRY_BASE_URL"]
except (FileNotFoundError, KeyError, AttributeError):
    AZURE_FOUNDRY_BASE_URL = os.getenv("AZURE_FOUNDRY_BASE_URL")

try:
    XAI_api_key = st.secrets["XAI_api_key"]
except (FileNotFoundError, KeyError, AttributeError):
    XAI_api_key = os.getenv("XAI_api_key")
    

# Cloud storage URL for production vector database (optional)
# Use HuggingFace Hub: huggingface_hub.download() with repo_id
try:
    HUGGINGFACE_AZURE_REPO_ID = st.secrets.get("HUGGINGFACE_AZURE_REPO")
except:
    HUGGINGFACE_AZURE_REPO_ID = os.getenv("HUGGINGFACE_AZURE_REPO")


# Keep instructions as a stable, shared prefix to improve provider-side token cache hit rates.
METADATA_PROMPT_TEMPLATE = """You are an expert in epidemiology and aging research, specializing in cohort study metadata. CRITICAL: Do NOT invent or hallucinate data. Only use information explicitly provided in NEAR metadata below.

COHORT BACKGROUND:
{cohort_background}

---

Your task is to answer questions about variables and metadata from this cohort.

### NEAR metadata (ONLY SOURCE OF TRUTH):
Each block below represents one retrieved variable definition.
Each block begins with a header in this format:
[Source: source_file_name | Variable: variable_name]
{context}

### Question from user:
{question}

### Your Response Instructions:
- Treat every question as a single-turn request based only on the current question and provided metadata context
- DO NOT include follow-up offers such as "If you want, I can...", "Would you like me to...", or "I can also..."
- Group related variables by theme
- Start with a clear, natural explanation of the topic based on the related cohort background
- Use your own words to describe what the variables measure
- Then, present the variable information in a markdown table with these columns:

| Variable Name | Label | Categories | Source |
|---|---|---|---|
| variable_name | What it measures in plain English | Category values if applicable | source_file_name |

================================================================
CRITICAL RULES FOR EXTRACTING DATA (READ CAREFULLY):
================================================================

1. VARIABLE NAMES (Column 1):
    - EXTRACT EXACTLY the text that appears after "Variable: " in the source data
    - Copy-paste the exact variable name - do NOT modify, shorten, or translate
    - Do NOT use field names like "Description" or "Label" instead
    - FORBIDDEN: Do NOT invent variable names not in the source
    - EXAMPLE RIGHT: Source has "Variable: löpnr". Write "löpnr" exactly
    - EXAMPLE WRONG: Inventing "participant_age" when source only has "löpnr"

2. LABELS (Column 2):
    - Use the "Label:" field value EXACTLY as written in source data
    - Do NOT shorten, paraphrase, or interpret the label

3. CATEGORIES (Column 3):
    - Extract category values EXACTLY as they appear in the source data, typically in a "Categories:" field
    - FORMAT STRICTLY as: "1=value1, 2=value2, 3=value3" (number=description, comma-separated, NO SEMICOLONS)
    - If no categories exist, write "N/A"
    - Do NOT invent category mappings not in the source

4. SOURCE (Column 4):
    - Extract from the header "[Source: filename | Variable: variable_name]" at the start of each block
    - Must be present for EVERY variable (required field)
    - Use the exact source filename provided

5. DEDUPLICATION:
    - Variables with identical Label and Categories are the same variable collected across different time points
    - COMBINES these into a single row with all variable names and sources listed
    - This provides visibility into data collection across waves while avoiding redundancy

6. VALIDATION (CRITICAL - MUST FOLLOW):
    - EVERY row must have ALL 4 columns completely filled (NO EMPTY CELLS)
    - NEVER pad rows with empty cells or partial data
    - Every variable name must come from the source data
    - If data is missing from source, DO NOT hallucinate it
    - Double-check: Each variable in your table should be traceable to the source context
    - If the last row is incomplete, DELETE IT and end the table cleanly

DETAILED EXAMPLES OF CORRECT FORMAT:

Example 1 - Demographics in SNAC-K:
"In SNAC-K, several variables measure basic demographics. Each participant has a unique identifier and recorded biological sex."

| Variable Name | Label | Categories | Source |
|---|---|---|---|
| löpnr | Unique participant identifier number | N/A | SNAC_K_Baseline |
| kön | Participant's biological sex | 1=man, 2=woman | SNAC_K_Baseline |

Example 2 - Physical Measurements:
"Height and weight are measured at baseline assessment."

| Variable Name | Label | Categories | Source |
|---|---|---|---|
| height_cm | Height in centimeters | N/A | H70_Baseline_Form |
| weight_kg | Body weight in kilograms | N/A | H70_Baseline_Form |

Answer:"""

METADATA_PROMPT = PromptTemplate(
     template=METADATA_PROMPT_TEMPLATE,
     input_variables=["cohort_background", "context", "question"],
)

# ── Functions ─────────────────────────────────────────────────────────────────

def _find_table_blocks(text):
    """Find all markdown table blocks in text.
    
    Args:
        text: String potentially containing markdown tables
    
    Returns:
        List of table blocks, where each block is a list of markdown table lines.
    """
    lines = text.split('\n')
    table_blocks = []
    current_table = []

    for original_line in lines:
        line = original_line.strip()

        if line.startswith('|') and line.endswith('|'):
            current_table.append(line)
            continue

        if current_table:
            table_blocks.append(current_table)
            current_table = []

    if current_table:
        table_blocks.append(current_table)

    return table_blocks



def _parse_markdown_table(table_lines):
    """Parse a list of markdown table lines into a DataFrame.
    
    Args:
        table_lines: List of markdown table lines (| col1 | col2 | ...)
    
    Returns:
        pandas DataFrame or None if invalid
    """
    if len(table_lines) < 2:
        return None
    
    try:
        # Parse header
        header_line = table_lines[0]
        headers = [cell.strip() for cell in header_line.split('|')[1:-1]]
        
        # Skip separator line (contains dashes)
        rows = []
        for line in table_lines[2:]:
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            
            # Skip incomplete rows entirely (don't pad with empty strings)
            # With improved LLM instructions, incomplete rows should be rare
            # If a row doesn't have all columns, skip it to maintain data integrity
            if len(cells) != len(headers):
                continue
            
            if cells:  # Only add non-empty rows
                rows.append(cells)
        
        if not rows:
            return None
        
        df = pd.DataFrame(rows, columns=headers)
        
        if df.empty:
            return None
        
        return df
    except Exception as e:
        return None
    
def extract_markdown_tables(text):
    """Extract markdown tables from text.
    
    Args:
        text: String containing markdown tables
    
    Returns:
        List of pandas DataFrames, one per table found
    """
    tables = []

    for table_lines in _find_table_blocks(text):
        df = _parse_markdown_table(table_lines)
        if df is not None and not df.empty:
            tables.append(df)

    return tables



def export_tables_to_excel(tables):
    """Export multiple DataFrames to Excel file with formatted sheets.
    
    Args: 
        tables: List of DataFrames
    
    Returns:
        BytesIO object containing the Excel file
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for idx, df in enumerate(tables):
            # Use simple deterministic sheet names to avoid fragile header parsing.
            sheet_name = f"Table {idx + 1}"
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the sheet
            worksheet = writer.sheets[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color="FF7BA74F", end_color="FF7BA74F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply borders to all cells
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border
                    if cell.row > 1:  # Data rows
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    output.seek(0)
    return output


def initialize_production_db(chroma_db: str, hf_repo_id: str, zip_filename: str):
    """Download production database from HuggingFace Hub if not present locally.
    
    This function:
    1. Checks if the database already exists locally (skip download if so)
    2. Downloads the compressed database from HuggingFace Hub
    3. Extracts the archive to a temporary directory
    4. Verifies the SQLite database was extracted correctly
    5. Moves the database to the final location (CHROMA_DB)
    
    The downloaded database contains pre-built Chroma vector collections
    for all cohort datasets (Betula, SNAC-K, etc.).
    """
    abs_chroma_db = os.path.abspath(chroma_db)
    db_parent = os.path.dirname(abs_chroma_db) or "."
    lock_dir = f"{abs_chroma_db}.init.lock"

    # ─ Check if database already exists locally ───────────────────────────────
    # Skip download if database directory exists and contains files
    # This prevents unnecessary downloads on subsequent app restarts
    if os.path.exists(abs_chroma_db) and os.listdir(abs_chroma_db):
        return

    # ─ Acquire a simple process-safe lock ─────────────────────────────────────
    # Streamlit can trigger concurrent reruns; prevent overlapping installs.
    lock_acquired = False
    lock_timeout_seconds = 120
    lock_start = time.time()
    while not lock_acquired:
        try:
            os.mkdir(lock_dir)
            lock_acquired = True
        except FileExistsError:
            # Another process/session is currently installing the DB.
            if os.path.exists(abs_chroma_db) and os.listdir(abs_chroma_db):
                return
            if (time.time() - lock_start) > lock_timeout_seconds:
                st.error("❌ Timed out waiting for database initialization lock.")
                st.stop()
            time.sleep(0.5)
    
    try:
        # ─ Validate required configuration ─────────────────────────────────────
        # Stop immediately if HuggingFace repository ID is not configured
        # This is a required secret in Streamlit Cloud
        if not hf_repo_id:
            st.error("❌ HUGGINGFACE_AZURE_REPO not configured. Please contact the maintainer.")
            st.stop()

        from huggingface_hub import hf_hub_download

        # ─ Show download progress to user ──────────────────────────────────────
        # Create a placeholder that we'll update as the download progresses
        progress_placeholder = st.empty()
        progress_placeholder.info("📥 Downloading production database from HuggingFace Hub...")

        # ─ Download the database zip file ──────────────────────────────────────
        # Downloads the configured Chroma database zip from HuggingFace dataset
        # File is cached in ~/.cache/huggingface to avoid re-downloading
        try:
            zip_path = hf_hub_download(
                repo_id=hf_repo_id,
                filename=zip_filename,
                repo_type="dataset",
                cache_dir=Path.home() / ".cache" / "huggingface"
            )
            logger.info(f"Downloaded database from HuggingFace: {hf_repo_id}")
        except Exception as e:
            # Handle download errors (network issues, missing file, auth errors, etc.)
            error_msg = f"Failed to download from HuggingFace: {e}"
            logger.error(error_msg)
            progress_placeholder.error(f"❌ {error_msg}")
            st.stop()

        # ─ Extract the downloaded archive ──────────────────────────────────────
        # Update progress message and create temporary directory for extraction
        progress_placeholder.info("📦 Extracting database...")
        temp_dir = tempfile.mkdtemp()

        try:
            # Extract all files from the zip archive
            with zipfile.ZipFile(zip_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)

            extracted_contents = os.listdir(temp_dir)

            # ─ Handle different zip structures ─────────────────────────────────
            # The zip file might contain either:
            # A) Direct contents: chroma.sqlite3, UUID folders, etc. (typical case)
            # B) Wrapped: A single "chroma_azure_db" folder containing everything
            #
            # We need to handle both cases to find the correct source path
            if "chroma_azure_db" in extracted_contents and len(extracted_contents) == 1:
                # Case B: Zip had a top-level chroma_azure_db wrapper folder
                source_path = os.path.join(temp_dir, "chroma_azure_db")
            else:
                # Case A: Zip had direct contents (most common)
                source_path = temp_dir

            # ─ Verify database integrity ──────────────────────────────────────
            # The SQLite database file (chroma.sqlite3) must exist
            # If missing, the zip was corrupted or incompatible
            sqlite_path = os.path.join(source_path, "chroma.sqlite3")
            if not os.path.exists(sqlite_path):
                error_files = os.listdir(source_path)[:5]
                progress_placeholder.error(f"❌ Database files corrupted. Found: {error_files}")
                st.stop()

            # ─ Clean up and copy to final location via staging ─────────────────
            # Avoid direct cross-filesystem moves from /tmp and ensure destination
            # parent directories exist before final swap.
            os.makedirs(db_parent, exist_ok=True)
            staging_path = f"{abs_chroma_db}.staging"

            if os.path.exists(staging_path):
                shutil.rmtree(staging_path)

            shutil.copytree(source_path, staging_path)

            if not os.path.exists(os.path.join(staging_path, "chroma.sqlite3")):
                progress_placeholder.error("❌ Staging verification failed (missing chroma.sqlite3).")
                st.stop()

            if os.path.exists(abs_chroma_db):
                shutil.rmtree(abs_chroma_db)

            shutil.move(staging_path, abs_chroma_db)

            # ─ Final verification ─────────────────────────────────────────────
            # Confirm the move was successful by checking for the SQLite file
            if os.path.exists(os.path.join(abs_chroma_db, "chroma.sqlite3")):
                logger.info("Database extracted and ready")
                progress_placeholder.success("✅ Database downloaded and ready!")
            else:
                # Move succeeded but database file is missing (should never happen)
                progress_placeholder.error("❌ Database extraction failed")
                st.stop()

        except Exception as e:
            # Handle zip extraction or file operation errors
            logger.error(f"Extraction failed: {e}")
            progress_placeholder.error(f"❌ Failed to extract: {e}")
            st.stop()
        finally:
            # ─ Cleanup temporary directory ─────────────────────────────────────
            # Always remove temp directory, even if an error occurred
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

    except ImportError:
        # Handle missing huggingface_hub package
        st.error("❌ Install huggingface_hub: `pip install huggingface_hub`")
        st.stop()
    except Exception as e:
        # Catch any other unexpected errors
        logger.error(f"Unexpected error: {e}")
        st.error(f"❌ {e}")
        st.stop()
    finally:
        if os.path.exists(lock_dir):
            shutil.rmtree(lock_dir)

@st.cache_resource(show_spinner="Preparing embeddings model...")
def get_embeddings():
    return AzureOpenAIEmbeddings(
        azure_deployment=EMBEDDING_MODEL_NAME,
        azure_endpoint=AZURE_openai_endpoint,
        api_key=AZURE_api_key,
        api_version=AZURE_EMBEDDING_API_VERSION,
    )


def get_available_databases(chroma_db: str):
    """Get list of available databases by querying Chroma collections."""
    databases = []
    
    try:
        if not os.path.exists(chroma_db):
            st.error(f"❌ Database directory not found: {chroma_db}")
            st.stop()
        
        if not os.listdir(chroma_db):
            st.error(f"❌ Database directory is empty: {chroma_db}")
            st.stop()
        
        embeddings = get_embeddings()
        
        for db_name in KNOWN_DATABASES:
            collection_name = f"{db_name.lower()}_metadata"
            try:
                vectorstore = Chroma(
                    persist_directory=chroma_db,
                    embedding_function=embeddings,
                    collection_name=collection_name
                )
                if vectorstore._collection.count() > 0:
                    databases.append(db_name)
            except Exception as e:
                logger.debug(f"Collection {db_name} not available: {e}")
        
        if not databases:
            logger.warning("No databases could be loaded from Chroma")
        else:
            logger.info(f"Loaded {len(databases)} databases")
        
        return sorted(databases)
    except Exception as e:
        logger.error(f"Error retrieving databases: {e}")
        st.error(f"❌ Could not load databases: {e}")
        return []

# Initialize vectorstore as None (will be set from cache on database selection)
if "vectorstore" not in st.session_state:
    st.session_state.vectorstore = None
    st.session_state.selected_database = None

# Initialize vectorstores cache (preload all databases on first access)
if "vectorstores_cache" not in st.session_state:
    st.session_state.vectorstores_cache = {}
    st.session_state.vectorstores_loading = False

# Initialize selected LLM model
if "selected_llm_model" not in st.session_state:
    st.session_state.selected_llm_model = LLM_MODEL_GROQ

# Initialize Azure GPT rate-limit counter (resets when the browser session resets)
if "gpt_query_count" not in st.session_state:
    st.session_state.gpt_query_count = 0

def get_complete_background(vectorstore):
    """Retrieve all database description chunks via semantic search.
    
    Uses semantic search to find and return all chunks tagged as database descriptions,
    providing comprehensive cohort background information as a fallback when query-specific
    background retrieval yields no results.
    """
    if vectorstore is None:
        return ""
    
    try:
        retriever = vectorstore.as_retriever(
            search_kwargs={
                "k": 2,
                "filter": {"type": "database_description"}
            }
        )
        docs = retriever.invoke("database description overview cohort study information")
        
        descriptions = [doc.page_content for doc in docs]
        
        return "\n\n".join(descriptions) if descriptions else ""
    except Exception as e:
        logger.error(f"Could not retrieve database description: {e}")
        return ""

def get_relevant_background(query, vectorstore, context_docs=None):
    """Retrieve cohort background information relevant to the user's query."""
    if vectorstore is None:
        return ""
    
    try:
        # Perform semantic search on the entire collection
        retriever = vectorstore.as_retriever(
            search_kwargs={
                "k": RETRIEVAL_K_BACKGROUND,
                "filter": {"type": "database_description"}
            }
        )
        background_query = query

        # If the user query looks like an exact variable code, enrich background
        # retrieval with matched variable labels/sources to improve relevance.
        candidates = extract_variable_name_candidates(query)
        if candidates and context_docs:
            candidate_set = {c.lower() for c in candidates}
            matched_docs = []

            for doc in context_docs:
                variable_name = (doc.metadata or {}).get("variable_name")
                if variable_name and variable_name.lower() in candidate_set:
                    matched_docs.append(doc)

            if not matched_docs:
                matched_docs = context_docs[:BACKGROUND_ENRICH_FALLBACK_DOCS]

            enrichment_terms = []
            seen_terms = set()
            for doc in matched_docs[:BACKGROUND_ENRICH_MAX_DOCS]:
                source = (doc.metadata or {}).get("source")
                if source:
                    clean_source = source.replace(".xml", "")
                    if clean_source.lower() not in seen_terms:
                        seen_terms.add(clean_source.lower())
                        enrichment_terms.append(clean_source)

                label_match = re.search(r"(?im)^Label:\s*(.+)$", doc.page_content or "")
                if label_match:
                    label = label_match.group(1).strip()
                    if label.lower() not in seen_terms:
                        seen_terms.add(label.lower())
                        enrichment_terms.append(label)

            if enrichment_terms:
                background_query = f"{query} {' '.join(enrichment_terms)}"

        docs = retriever.invoke(background_query)
        
        background_docs = docs
        
        if not background_docs:
            return get_complete_background(vectorstore)
        
        return "\n\n---\n\n".join(
            [doc.page_content for doc in background_docs]
        )
    except Exception as e:
        logger.error(f"Error retrieving background: {e}")
        return get_complete_background(vectorstore)



def extract_variable_name_candidates(query: str) -> list[str]:
    """Extract likely variable identifiers from a user query.

    Targets code-like identifiers such as BAS1_H4, v518, or quoted variable names.
    """
    if not query:
        return []

    candidates = []
    seen = set()

    def add_candidate(value: str) -> None:
        cleaned = value.strip().strip("`'\"")
        if not cleaned:
            return
        if cleaned.lower() in seen:
            return
        seen.add(cleaned.lower())
        candidates.append(cleaned)

    for match in re.findall(r"[`\"']([A-Za-z][A-Za-z0-9_:-]*)[`\"']", query):
        add_candidate(match)

    for token in re.findall(r"\b[A-Za-z][A-Za-z0-9_:-]*\b", query):
        if any(char.isdigit() for char in token) or "_" in token or "-" in token or token.isupper():
            add_candidate(token)

    return candidates


def _collection_payload_to_documents(payload: dict | None) -> list[Document]:
    """Convert a Chroma collection.get payload into LangChain Documents."""
    if not payload:
        return []

    documents = payload.get("documents") or []
    metadatas = payload.get("metadatas") or []
    results = []

    for index, page_content in enumerate(documents):
        metadata = metadatas[index] if index < len(metadatas) else {}
        results.append(Document(page_content=page_content, metadata=metadata or {}))

    return results


def deduplicate_documents(documents: list[Document]) -> list[Document]:
    """Deduplicate retrieval results while preserving order."""
    unique_docs = []
    seen = set()

    for doc in documents:
        metadata = doc.metadata or {}
        key = (
            metadata.get("source"),
            metadata.get("variable_name"),
            doc.page_content,
        )
        if key in seen:
            continue
        seen.add(key)
        unique_docs.append(doc)

    return unique_docs


def format_context_for_prompt(documents: list[Document]) -> str:
    """Format retrieval docs into a deterministic context block for prompting."""
    # Stable ordering keeps prompt prefixes more consistent across similar queries.
    sorted_docs = sorted(
        documents,
        key=lambda doc: (
            ((doc.metadata or {}).get("source") or "").lower(),
            ((doc.metadata or {}).get("variable_name") or "").lower(),
            (doc.page_content or ""),
        ),
    )

    context_parts = []
    for doc in sorted_docs:
        source = (doc.metadata or {}).get("source", "Unknown")
        variable_name = (doc.metadata or {}).get("variable_name")

        if source.endswith(".xml"):
            source = source[:-4]

        if variable_name:
            context_parts.append(f"[Source: {source} | Variable: {variable_name}]\n{doc.page_content}")
        else:
            context_parts.append(f"[Source: {source}]\n{doc.page_content}")

    return "\n\n---\n\n".join(context_parts)


def exact_variable_name_lookup(query: str, vectorstore) -> list[Document]:
    """Look up variable documents by exact metadata match before semantic search."""
    if vectorstore is None:
        return []

    collection = getattr(vectorstore, "_collection", None)
    if collection is None:
        return []

    candidates = extract_variable_name_candidates(query)
    if not candidates:
        return []

    exact_docs = []

    for candidate in candidates:
        candidate_variants = [candidate]
        for variant in (candidate.upper(), candidate.lower()):
            if variant not in candidate_variants:
                candidate_variants.append(variant)

        for variant in candidate_variants:
            payload = collection.get(
                where={"variable_name": variant},
                include=["documents", "metadatas"],
                limit=EXACT_MATCH_LIMIT,
            )
            exact_docs.extend(_collection_payload_to_documents(payload))

        if exact_docs:
            continue

        payload = collection.get(
            where={"type": "variable_definitions"},
            where_document={"$contains": f"Variable: {candidate}"},
            include=["documents", "metadatas"],
            limit=EXACT_MATCH_LIMIT,
        )
        exact_docs.extend(_collection_payload_to_documents(payload))

    exact_docs = deduplicate_documents(exact_docs)
    if exact_docs:
        logger.debug(
            "Exact variable lookup matched %s docs for candidates %s",
            len(exact_docs),
            candidates,
        )

    return exact_docs

def filter_and_organize_context(query, vectorstore, llm_model=None):
    """Retrieve variable definitions relevant to the query.
    
    Returns context with source information for each variable.
    """
    if vectorstore is None:
        return "", []
    
    try:
        # Adjust retrieval depth based on LLM model.
        k_context = RETRIEVAL_K_CONTEXT_GROQ if llm_model == LLM_MODEL_GROQ else RETRIEVAL_K_CONTEXT
        retriever = vectorstore.as_retriever(
            search_kwargs={
                "k": k_context,
                "filter": {"type": "variable_definitions"}
            }
        )
        docs = retriever.invoke(query)
        exact_docs = exact_variable_name_lookup(query, vectorstore)
        var_defs = deduplicate_documents(exact_docs + docs)
        
        # Debug logging: show how many documents were retrieved
        logger.debug(
            f"Query: '{query}' | k_context: {k_context} "
            f"| Semantic docs: {len(docs)} | Exact docs: {len(exact_docs)} "
            f"| Final docs: {len(var_defs)} | Model: {llm_model}"
        )
        
        context_text = format_context_for_prompt(var_defs)
        logger.debug(f"Final context: {len(var_defs)} variable definitions included ({len(context_text)} chars)")
        return context_text, var_defs
    except Exception as e:
        error_msg = f"Error retrieving context: {e}"
        logger.error(error_msg)
        logger.exception("Full traceback:")
        # Note: Don't show error in UI here since this is called during response generation
        return "", []

def get_llm(model_name: str):
    """Initialize the selected LLM model.
    
    Args:
        model_name: Name of the model to initialize
    
    Returns:
        Initialized LLM instance
    """
    if model_name == LLM_MODEL_GPT:
        if not AZURE_api_key:
            st.error("❌ AZURE_api_key not configured")
            st.stop()
        return ChatOpenAI(
            api_key=AZURE_api_key,
            model=GPT_MODEL_ID,
            base_url=AZURE_FOUNDRY_BASE_URL,
            temperature=LLM_TEMPERATURE,
        )
    elif model_name == LLM_MODEL_XAI_GROK:
        if not XAI_api_key:
            st.error("❌ XAI_api_key not configured")
            st.stop()
        return ChatOpenAI(
            api_key=XAI_api_key,
            model=XAI_GROK_MODEL_ID,
            base_url=XAI_BASE_URL,
            temperature=LLM_TEMPERATURE,
        )
    else:  # Default to Groq Llama 3.1 8B
        if not GROQ_API_KEY:
            st.error("❌ GROQ_api_key not configured")
            st.stop()
        return ChatGroq(
            groq_api_key=GROQ_API_KEY,
            model_name=GROQ_MODEL_ID,
            temperature=LLM_TEMPERATURE,
        )




# ── Main App ──────────────────────────────────────────────────────────────────

st.title("NEAR Metadata Chatbot")

# Verify HUGGINGFACE_AZURE_REPO is available before initializing database
if not HUGGINGFACE_AZURE_REPO_ID:
    st.error(f"❌ {HUGGINGFACE_REPO_SECRET_NAME} not configured in Streamlit secrets.")
    st.stop()

# Initialize production database from cloud if needed (MUST BE FIRST!)
initialize_production_db(CHROMA_DB, HUGGINGFACE_AZURE_REPO_ID, CHROMA_ZIP_FILENAME)

# Initialize available databases (after database is ready)
if not st.session_state.get("available_databases_loaded", False):
    with st.spinner("Discovering available databases..."):
        st.session_state.available_databases = get_available_databases(CHROMA_DB)
        st.session_state.available_databases_loaded = True
    
    if not st.session_state.available_databases:
        st.error("""
        ❌ **No databases were discovered!**
        
        Please try:
        - Click "Refresh" in the sidebar to retry
        - Contact the maintainer
        """)

# Preload all vectorstores in the background (cache them)
if not st.session_state.vectorstores_loading and st.session_state.available_databases:
    if len(st.session_state.vectorstores_cache) == 0:
        st.session_state.vectorstores_loading = True
        with st.spinner("Preloading vector stores..."):
            embeddings = get_embeddings()
            for db in st.session_state.available_databases:
                try:
                    collection_name = f"{db.lower()}_metadata"
                    vectorstore = Chroma(
                        persist_directory=CHROMA_DB,
                        embedding_function=embeddings,
                        collection_name=collection_name
                    )
                    if vectorstore._collection.count() > 0:
                        st.session_state.vectorstores_cache[db] = vectorstore
                except Exception as e:
                    st.warning(f"Could not preload {db}: {e}")
        st.session_state.vectorstores_loading = False

# Add controls to sidebar
with st.sidebar:
    logo_path = Path("logo/NEAR-chatbot.jpg")
    if logo_path.exists():
        col1, col2, col3 = st.columns([0.5, 2.5, 0.5])
        with col2:
            st.image(str(logo_path), width='stretch')
    st.markdown("---")
    
    # Database selection
    st.subheader("Select Database")
    
    if st.session_state.available_databases:
        selected_database = st.radio(
            "Choose a database to query:",
            options=sorted(st.session_state.available_databases),
            index=0,
            key="database_radio"
        )

        # Always bind the selected DB to the current embedding-model cache.
        # This avoids stale state on the first query after switching embeddings.
        if selected_database in st.session_state.vectorstores_cache:
            st.session_state.vectorstore = st.session_state.vectorstores_cache[selected_database]
            st.session_state.selected_database = selected_database
        else:
            st.error(f"Vector store for {selected_database} not available")
        
    else:
        st.warning("No databases available")
    
    st.markdown("---")
    
    # LLM Model Selection
    st.subheader("Select LLM Model")
    available_models = [LLM_MODEL_GROQ, LLM_MODEL_GPT, LLM_MODEL_XAI_GROK]
    selected_model = st.radio(
        "Choose LLM model:",
        options=available_models,
        index=0,
        key="llm_radio"
    )
    st.session_state.selected_llm_model = selected_model
    
    if selected_model == LLM_MODEL_GPT:
        st.caption("🧠 GPT-5.4 Mini for stronger reasoning on metadata questions")
    elif selected_model == LLM_MODEL_XAI_GROK:
        st.caption("🚀 Fast reasoning with xAI Grok 4.1")
    else:
        st.caption("⚡ Faster responses with slightly smaller context window (may miss some category info)")

    st.markdown("---")
    
    # Session usage display (Azure GPT only)
    st.subheader("Session Usage")
    gpt_used = st.session_state.gpt_query_count
    gpt_left = max(0, RATE_LIMIT_MAX_GPT_QUERIES - gpt_used)
    st.caption(f"Azure GPT queries: {gpt_used} / {RATE_LIMIT_MAX_GPT_QUERIES} ({gpt_left} remaining)")
    if gpt_left == 0:
        st.warning(f"Azure GPT limit reached. Switch to {LLM_MODEL_GROQ} or start a new session.")
    elif gpt_left <= 3:
        st.caption(f"⚠️ Only {gpt_left} Azure GPT queries left this session.")

    st.markdown("---")
    
    # Contact & Support
    st.subheader("Contact & Support")
    st.markdown("""
    **Maintainer:** [Bolin Wu (NEAR)](https://ki.se/en/people/bolin-wu)
    
    **E-mail:** [📧 bolin.wu@ki.se](mailto:bolin.wu@ki.se)
    """)
    
    st.markdown("---")
    
    # Changelog
    with st.expander("📋 What's New"):
        st.markdown("""
        **v1.4** – Azure Embedding Upgrade
        - Switched embedding model to `text-embedding-3-small` (Azure OpenAI)
        - Retrieval now uses Azure embedding pipeline
        - Expected better semantic matching for metadata questions (stronger query-to-variable alignment)
        
        [View all previous releases →](https://github.com/Bolin-Wu/NEAR-metadata-chatbot/releases)
        """)

# Add disclaimer and tip about reference information
st.info("""
**ℹ️ Disclaimer:** This chatbot provides a quick overview based on NEAR metadata submitted to Maelstrom. Some variables may be unavailable, including derived variables, registry data, or recent updates. For complete and validated metadata, refer to the [Maelstrom catalogue](https://www.maelstrom-research.org/search#lists?type=studies&query=network(in(Mica_network.id,near)),variable(limit(0,20)),study(in(Mica_study.className,Study),limit(0,20))) or contact the NEAR team.

**💡 Tip:** If you're not satisfied with the results, try searching again with different wording or switch AI models. The same question may yield different results due to the nature of AI-powered responses.
""")

# Display search suggestions
st.markdown("### Example prompts:")
col1, col2, col3 = st.columns(3)
with col1:
    st.caption("- What cognitive tests are included?")
    st.caption("- How is sleep measured?")
    st.caption("- What social engagement data do you have?")
with col2:
    st.caption("- What nutrition and dietary related variables are available?")
    st.caption("- Recommend variables for constructing CIRS.")
    st.caption("- Suggest variables for frailty index.")
with col3:
    st.caption("- How do you assess mental health?")
    st.caption("- List ADL variables and map each to core ADL, and IADL.")
    st.caption("- Give all bathing, dressing, toileting, transferring variables with source and categories.")


# ── Chat ──────────────────────────────────────────────────────────────────────
if "messages" not in st.session_state:
    st.session_state.messages = []

for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

placeholder_text = f"Ask about {st.session_state.selected_database} metadata..."
if prompt := st.chat_input(placeholder_text):
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    with st.chat_message("assistant"):
        vectorstore = st.session_state.get("vectorstore")
        if vectorstore is None:
            st.warning("Vector store not available!")
            response = ""
        else:
            with st.spinner("Thinking..."):
                # ── Rate-limit check (Azure GPT only) ────────────────────────
                chosen_model = st.session_state.selected_llm_model
                gpt_limit_hit = (
                    chosen_model == LLM_MODEL_GPT
                    and st.session_state.gpt_query_count >= RATE_LIMIT_MAX_GPT_QUERIES
                )

                if gpt_limit_hit:
                    st.warning(
                        f"⚠️ Azure GPT limit reached for this session ({RATE_LIMIT_MAX_GPT_QUERIES} queries). "
                        f"Switch to **{LLM_MODEL_GROQ}** or start a new session."
                    )
                    response = ""
                else:
                    # Increment counter before the call so a crashed call still counts
                    if chosen_model == LLM_MODEL_GPT:
                        st.session_state.gpt_query_count += 1

                    # Get the selected LLM model
                    llm = get_llm(st.session_state.selected_llm_model)
                    
                    # Display which model is being used
                    st.caption(f"🔧 Using: {st.session_state.selected_llm_model} | Embedding: {EMBEDDING_MODEL_NAME}")
                    
                    # Get context (already filtered to selected database via collection)
                    # Returns both context text and document list
                    context, context_docs = filter_and_organize_context(prompt, vectorstore, st.session_state.selected_llm_model)

                    # Avoid blank assistant output when retrieval returns no variable chunks.
                    if not context.strip():
                        response = (
                            "I could not find matching variable definitions for that query in the selected database. "
                            "Try a more specific term (for example: systolic, diastolic, hypertension, blood pressure)."
                        )
                        logger.warning(
                            "Empty retrieval context for query '%s' in database '%s' using model '%s'",
                            prompt,
                            st.session_state.get("selected_database"),
                            st.session_state.selected_llm_model,
                        )
                    else:

                        rag_chain = (
                            {
                                "cohort_background": lambda user_query: get_relevant_background(user_query, vectorstore, context_docs),
                                "context": lambda _: context,
                                "question": RunnablePassthrough()
                            }
                                     | METADATA_PROMPT
                            | llm
                            | StrOutputParser()
                        )

                        try:
                            logger.debug(f"Starting LLM invocation with query: {prompt[:100]}...")
                            logger.debug(f"Context length: {len(context)} chars, {len(context_docs)} docs")
                            logger.debug(f"Using model: {st.session_state.selected_llm_model}")
                            response = rag_chain.invoke(prompt)
                            logger.debug(f"LLM Response received (first 500 chars): {response[:500]}")
                        except Exception as e:
                            error_str = str(e)
                            logger.error(f"LLM Error: {error_str}")
                            logger.exception("Full exception traceback:")
                            
                            # Check for token limit exceeded error
                            current_model = st.session_state.selected_llm_model
                            if "rate_limit_exceeded" in error_str and "tokens" in error_str.lower():
                                st.error(f"⚠️ Request too large for {current_model}. Try asking about a smaller subset of variables.")
                                response = ""
                            elif "rate_limit" in str(e).lower() or "429" in str(e):
                                st.error(f"⚠️ Rate limit reached for {current_model}. Please try again in a few moments.")
                                response = ""
                            else:
                                st.error(f"Error: {str(e)}")
                                response = (
                                    "I ran into an issue while generating the answer. "
                                    "Please try again, switch model, or narrow the query."
                                )

        # Prepend database hint to response
        selected_db = st.session_state.get("selected_database")
        
        # Trim whitespace while preserving the original LLM response text.
        if response:
            response = response.strip()

        # Some models may return empty/whitespace output for no-hit queries.
        # Ensure users always receive an explicit fallback message.
        if not response:
            response = (
                "I could not find directly related variable definitions for that query in the selected database. "
                "Try a different keyword or a broader phrasing."
            )
        
        if selected_db and response:
            response_with_hint = f"📍 **{selected_db}**\n\n{response}"
        else:
            response_with_hint = response
        
        st.markdown(response_with_hint)
        
        # Extract tables from response and enable download if found
        tables = extract_markdown_tables(response)
        
        # Add download button if tables were found
        if tables:
            excel_file = export_tables_to_excel(tables)
            st.download_button(
                label=f"📥 Download as Excel ({len(tables)} table{'s' if len(tables) > 1 else ''})",
                data=excel_file,
                file_name=f"NEARchatbot_{st.session_state.selected_database}_{pd.Timestamp.now().strftime('%y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_tables"
            )

    st.session_state.messages.append({"role": "assistant", "content": response_with_hint})

if st.button("Clear Chat"):
    st.session_state.messages = []
    st.rerun()